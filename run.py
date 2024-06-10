import asyncio
import json
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from ebay import scrape_search, scrape_product  # Ensure ebay.py contains these functions
import sys

if __name__ == "__main__":
    # Access the first argument (after the script name)
    if len(sys.argv) > 1:  
        argument1 = sys.argv[1]
        print("You provided the argument:", argument1)
    else:
        print("No argument was provided.")
output = Path(__file__).parent / "results"
output.mkdir(exist_ok=True)

class DateTimeEncoder(json.JSONEncoder):
    """Custom JSONEncoder subclass that knows how to encode datetime values."""
    def default(self, o):
        if isinstance(o, datetime):
            return o.isoformat()  # Convert datetime objects to ISO-8601 string format
        return super(DateTimeEncoder, self).default(o)  # Default behaviour for other types

def convert_price_to_int(data):
    """Convert price fields to integers if possible."""
    for item in data:
        if isinstance(item, dict):
            for key in item.keys():
                if 'price' in key and item[key] is not None:
                    try:
                        item[key] = int(float(item[key]))
                    except ValueError:
                        pass  # If conversion fails, keep the original value
    return data

def add_table_to_sheet(sheet, table_name):
    """Add a table to the given sheet with a unique table name."""
    table_range = f"A1:{chr(65 + sheet.max_column - 1)}{sheet.max_row}"
    tab = Table(displayName=table_name, ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    tab.tableStyleInfo = style
    sheet.add_table(tab)

# Initialize a global index variable
global_index = 0

async def run():
    global global_index  # Access the global index variable

    print("Running eBay.com scrape and saving results to ./results directory")

    search_url = "https://www.ebay.com/sch/i.html?_from=R40&_trksid=p4432023.m570.l1312&_nkw="
    search_url += sys.argv[1]
    search_url += "&_sacat=0"
    search_results = await scrape_search(search_url, sys.argv[2], max_pages=2, start_index=global_index)

    # Update the global index after scraping
    global_index += len(search_results)

    # Convert search results to DataFrame and save as Excel
    df_search = pd.DataFrame(search_results)
    excel_path = output / "ebay_data.xlsx"
    with pd.ExcelWriter(excel_path, engine='openpyxl') as excel_writer:
        df_search.to_excel(excel_writer, sheet_name='Search Results', index=False)
        
        # Scrape single product and add description
        single_product_result = await scrape_product("https://www.ebay.com/itm/332562282948", sys.argv[2])
        single_product_result = convert_price_to_int([single_product_result])  # Convert prices to integers
        df_single_product = pd.DataFrame(single_product_result)
        df_single_product.to_excel(excel_writer, sheet_name='Single Product', index=False)

        # Scrape product variants and add descriptions
        variant_product_result = await scrape_product("https://www.ebay.com/itm/393531906094", sys.argv[2])
        if isinstance(variant_product_result, dict):
            variant_product_result = [variant_product_result]
        variant_product_result = convert_price_to_int(variant_product_result)  # Convert prices to integers
        df_variant_product = pd.DataFrame(variant_product_result)
        df_variant_product.to_excel(excel_writer, sheet_name='Product Variants', index=False)

    # Load the workbook and add tables
    workbook = load_workbook(excel_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        table_name = f"Table_{sheet_name.replace(' ', '_')}"
        add_table_to_sheet(sheet, table_name)
    workbook.save(excel_path)

if __name__ == "__main__":
    asyncio.run(run())
